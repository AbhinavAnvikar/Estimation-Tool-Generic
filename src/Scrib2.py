import openpyxl as xl

wb = xl.load_workbook('C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimation.xlsx')
sheet = wb['Estimation_LL']

print(sheet.max_row)
print(sheet.max_column)

#for i in range(1,sheet.max_row):

def get_LOE_LL(com,tab,trans,valid):
    #get the weights from estimation matrix based on the complexity

    for rows in range(1, sheet.max_row + 1):
        #print(sheet.cell(rows, 1).value)
        if sheet.cell(rows, 1).value == com:
            #print("reached here")
            for headers in range(1, sheet.max_column + 1):
                if sheet.cell(1, headers).value == 'Weights':
                    w_com = float(sheet.cell(rows, headers).value)
                elif sheet.cell(1, headers).value == 'Table':
                    w_tab = float(sheet.cell(rows, headers).value)
                elif sheet.cell(1, headers).value == 'Trans':
                    w_trans = float(sheet.cell(rows, headers).value)
                elif sheet.cell(1, headers).value == 'Valid':
                    w_val = float(sheet.cell(rows, headers).value)
                elif sheet.cell(1, headers).value == 'UT':
                    w_UT = float(sheet.cell(rows, headers).value)

    # calculate the LOE based on the matrix

    Sum_LOE = (((tab * w_tab) + ((trans * w_trans) + (valid * w_val)))* w_com)
    Toatl_loe = (Sum_LOE * (1+w_UT/100))

    return Sum_LOE,Toatl_loe

# loe,final_loe = get_LOE('Simple',1,1,1)
# print(loe,final_loe)
loe,final_loe = get_LOE('Complex',3,4,4)
print(loe,final_loe)

#
# for rows in range(1, sheet.max_row+1):
#     print(sheet.cell(rows,1).value)
#     if sheet.cell(rows,1).value == 'Simple':
#         print("reached here")
#         for headers in range(1,sheet.max_column + 1):
#             if sheet.cell(1,headers).value == 'Weights':
#                 print(sheet.cell(rows,headers).value)
#             elif sheet.cell(1,headers).value == 'Table':
#                 print(sheet.cell(rows,headers).value)
#             elif sheet.cell(1,headers).value == 'Trans':
#                 print(sheet.cell(rows,headers).value)
#             elif sheet.cell(1,headers).value == 'Valid':
#                 print(sheet.cell(rows,headers).value)
#             elif sheet.cell(1,headers).value == 'UT':
#                 print(sheet.cell(rows,headers).value)
#







#
# lst=[]
#
# for i in range(1,sheet.max_row):
#     if sheet.cell(i,1).value == 'Simple':   # com 1 value passed
#         lst.append(i)
#         print(i,sheet.cell(i,1).value)
#
# print(sheet.max_row)
# print(sheet.max_column)
# for rows in range(1, sheet.max_row + 1):
#     if sheet.cell(rows,1).value == 'Complex' and sheet.cell(rows,2).value == 5:
#         for headers in range(1,sheet.max_column + 1):
#             if sheet.cell(1,headers).value == 'Tech_LOE':
#                 print(sheet.cell(rows,headers).value)
#             elif sheet.cell(1,headers).value == 'Final_Tech_LOE':
#                 print(sheet.cell(rows,headers).value)

