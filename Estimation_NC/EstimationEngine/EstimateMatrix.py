import openpyxl as xl
import pandas as pd
#import datetime

class Estimate_Engine:
    @staticmethod
    def total_estimate(lst):
        sum_loe=0.00
        sum_loe_ut=0.00
        for item in lst:
            sum_loe+=item[3]
            sum_loe_ut += item[4]

        return round(sum_loe,2),round(sum_loe_ut,2)

    @staticmethod
    def total_estimate_ll(lst):
        sum_loe = 0.00
        sum_loe_ut = 0.00
        for item in lst:
            sum_loe += item[4]
            sum_loe_ut += item[5]

        return round(sum_loe, 2), round(sum_loe_ut, 2)

    @staticmethod
    def get_loe(com1,com2):
        tech_loe = 0.0
        final_tech_loe=0.0
        wb = xl.load_workbook("C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimation.xlsx")
        sheet = wb['Estimation_Matrix']
        for rows in range(1, sheet.max_row + 1):
            if sheet.cell(rows, 1).value == com1 and str(sheet.cell(rows, 2).value) == com2:
                for headers in range(1, sheet.max_column + 1):
                    if sheet.cell(1, headers).value == 'Tech_LOE':
                        tech_loe = sheet.cell(rows, headers).value
                    elif sheet.cell(1, headers).value == 'Final_Tech_LOE':
                        final_tech_loe = sheet.cell(rows, headers).value
                    if tech_loe and final_tech_loe:
                        break

        return tech_loe,final_tech_loe

    @staticmethod
    def get_loe_ll(com, tab, trans, valid):
        # get the weights from estimation matrix based on the complexity
        wb = xl.load_workbook(
            'C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimation.xlsx')
        sheet = wb['Estimation_LL']
        for rows in range(1, sheet.max_row + 1):
            # print(sheet.cell(rows, 1).value)
            if sheet.cell(rows, 1).value == com:
                print("reached here")
                for headers in range(1, sheet.max_column + 1):
                    if sheet.cell(1, headers).value == 'Weights':
                        w_com = float(sheet.cell(rows, headers).value)
                    elif sheet.cell(1, headers).value == 'Table':
                        w_tab = float(sheet.cell(rows, headers).value)
                    elif sheet.cell(1, headers).value == 'Trans':
                        w_trans = float(sheet.cell(rows, headers).value)
                    elif sheet.cell(1, headers).value == 'Valid':
                        w_val = float(sheet.cell(rows, headers).value)
                    elif sheet.cell(1, headers).value == 'UT':
                        w_UT = float(sheet.cell(rows, headers).value)

        # calculate the LOE based on the matrix

        Sum_LOE = (((float(tab) * w_tab) + ((float(trans) * w_trans) + (float(valid) * w_val))) * w_com)
        Total_loe = (Sum_LOE * (1 + w_UT / 100))
        return round(Sum_LOE,2), round(Total_loe,2)

    @staticmethod
    def download_to_excel(lst):
        print("point1")
        csv_list = [['Requirement', 'Complexity 1', 'Complexity 2', 'Tech_LOE', 'Tech_LOE_With_UT', 'Comments']]
        for i in range(0, len(lst)):
            csv_list.append(lst[i])
        # csv_list.append(rows)
        #str = str(datetime.datetime.now().time())
        #excel_name = "C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimate_Sheet_" + str
        writer = pd.ExcelWriter("C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimation_List.xlsx")
        df = pd.DataFrame(data=csv_list)
        df.to_excel(writer, header=False, index=False, sheet_name="Estimation_Sheet")
        writer.save()

    def download_to_excel_LL(lst):
        print("point1")
        csv_list = [['Number of Tables', 'Complexity', 'Avg Trans', 'Avg Valid', 'Tech_LOE(in MDs)','Tech_LOE_With_UT','Comments']]
        for i in range(0, len(lst)):
            csv_list.append(lst[i])
        # csv_list.append(rows)
        # str = str(datetime.datetime.now().time())
        # excel_name = "C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimate_Sheet_" + str
        writer = pd.ExcelWriter(
            "C:\\Users\\aban0617\\PycharmProjects\\Estimation_Tool\\venv\\Estimation_NC\\media\\Estimation_List_Low_Level.xlsx")
        df = pd.DataFrame(data=csv_list)
        df.to_excel(writer, header=False, index=False, sheet_name="Estimation_Sheet")
        writer.save()

# lst=[[0,0,0,1,2],[0,0,0,2,3]]
# print(lst[0][4])
# Estimate_Engine.total_estimate(lst)